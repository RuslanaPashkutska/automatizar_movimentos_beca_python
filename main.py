import pandas as pd
from difflib import SequenceMatcher
from openpyxl import load_workbook
from copy import copy


UMBRAL_SIMILITUD = 0.75
INPUT_FILE = "InputPL.xlsx"
MAYOR_FILE = "Mayor_TSCFO.xlsx"
OUTPUT_FILE = "InputPL_actualizado.xlsx"



#Limpieza
def clean_dataframe(df):
    df.columns = df.columns.str.strip()
    df = df[df["Fecha"].astype(str).str.upper() != "END"]
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"])
    df["Debe"] = pd.to_numeric(df["Debe"], errors="coerce").round(2)
    df["Haber"] = pd.to_numeric(df["Haber"], errors="coerce").round(2)
    df["Concepto"] = df["Concepto"].astype(str).str.strip().str.upper()
    return df


# Genera un identificador único combinando campos clave
# para detectar duplicados entre ambos archivos
def create_id(df):
    return (
        df["Fecha"].dt.strftime("%Y-%m-%d")
        + "_"
        + df["Nº Asiento"].astype(str)
        + "_"
        + df["Cuenta"].astype(str)
        + "_"
        + df["Debe"].round(2).astype(str)
        + "_"
        + df["Haber"].round(2).astype(str)
     )


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Busca el tipo de gasto más probable según similitud histórica
def assign_tipo(concepto, mapping):
    if concepto in mapping:
        return mapping[concepto], 1.0

    best_score = 0
    best_tipo = None

    for hist_concept, tipo in mapping.items():
        score = similarity(concepto, hist_concept)

        if score > best_score:
            best_score = score
            best_tipo = tipo

        if score == 1.0:
            break

    return best_tipo, best_score

def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

# Clasifica nuevos movimientos utilizando el histórico
# y calcula porcentaje de confianza
def classify_movements(new_movements, inputpl):
    mapping = (
        inputpl
        .dropna(subset=["Tipo de gasto"])
        .groupby("Concepto")["Tipo de gasto"]
        .agg(lambda x: x.mode().iloc[0])
        .to_dict()
    )
    tipos = []
    confianzas = []

    for _, row in new_movements.iterrows():
        tipo, score = assign_tipo(row["Concepto"], mapping)

        if score > UMBRAL_SIMILITUD:
            tipos.append(tipo)
        else:
            tipos.append("REVISAR")

        confianzas.append(round(score * 100, 2))

    new_movements["Tipo de gasto"] = tipos
    new_movements["Confianza"] = confianzas

    return new_movements


def main():
    inputpl = pd.read_excel(INPUT_FILE)
    mayor = pd.read_excel(MAYOR_FILE)

    mayor = mayor.rename(columns={"Net": "Neto", "Month": "Mes"})

    inputpl.columns = inputpl.columns.str.strip()
    mayor.columns = mayor.columns.str.strip()

    required_columns = [
        "Fecha", "Nº Asiento", "Concepto", "Cuenta", "Debe", "Haber", "Neto"
    ]
    for col in required_columns:
        if col not in inputpl.columns:
            raise ValueError(f"Falta de la columna '{col}' en InputPL")

        if col not in mayor.columns:
            raise ValueError(f"Falta de la columna '{col}' en Mayor_TSCFO")


    inputpl = clean_dataframe(inputpl)
    mayor = clean_dataframe(mayor)

    inputpl["ID"] = create_id(inputpl)
    mayor["ID"] = create_id(mayor)

    # Detecta movimientos nuevos
    new_movements = mayor[~mayor["ID"].isin(inputpl["ID"])].copy()
    new_movements = new_movements.sort_values("Fecha")

    if new_movements.empty:
        print("No se han detectado movimientos nuevos.")
        return

    new_movements = classify_movements(new_movements, inputpl)

    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # Buscar fila END
    end_row = None
    for r in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        if cell_value and str(cell_value).upper() == "END":
            end_row = r
            break

    if end_row is None:
        end_row = ws.max_row + 1

    columns_input = ["Nº Asiento", "Fecha", "Documento", "Concepto", "Cuenta",
        "Debe", "Haber", "Saldo", "Nombre cuenta", "Neto", "Mes", "Tipo de gasto"]

    # Encuentra última fila con datos para copiar formato
    last_data_row = end_row - 1
    while last_data_row > 0:
        if any(ws.cell(row=last_data_row, column=col).value for col in range(1, ws.max_column + 1)):
            break
        last_data_row -= 1

    # Insertar nuevas filas antes de END manteniendo formato
    for i, (_, row_data) in enumerate(new_movements.iterrows()):
        insert_position = end_row + i
        ws.insert_rows(insert_position)

        copy_row_format(ws, last_data_row, insert_position)

        for col_idx, column_name in enumerate(columns_input, 1):
            if column_name in row_data:
                valor = row_data[column_name]
                if column_name == "Fecha" and pd.notna(valor):
                    valor = valor.date()
                ws.cell(row=insert_position, column=col_idx, value=valor)

        print(f"Fila insertada: {row_data['Concepto']} -> {row_data['Tipo de gasto']}")


    wb.save(OUTPUT_FILE)

    total_nuevos =len(new_movements)
    clasificados = (new_movements["Confianza"] >= UMBRAL_SIMILITUD * 100).sum()
    revisar = total_nuevos - clasificados

    print(f"\nArchivo guardado como: {OUTPUT_FILE}")
    print("\nResumen de ejecución:")
    print(f"Movimientos nuevos detectados: {total_nuevos}")
    print(f"Clasificados automáticamente: {clasificados}")
    print(f"Requieren revisión manual: {revisar}")



if __name__ == "__main__":
    main()
